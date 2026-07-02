from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

import io

from sklearn.experimental import enable_iterative_imputer

try:
    from sklearn.impute import KNNImputer, IterativeImputer
    from sklearn.preprocessing import LabelEncoder
    SKLEARN_AVAILABLE = True
except ImportError:
    KNNImputer = None
    IterativeImputer = None
    LabelEncoder = None
    SKLEARN_AVAILABLE = False

# palette
C_HEADER_BG = '1E2330'
C_HEADER_FG = 'E8EAF0'
C_SUBHEADER = '2A3045'
C_ACCENT = '4F6EF7'
C_GREEN = '22C55E'
C_YELLOW = 'F59E0B'
C_RED = 'EF4444'
C_ROW_ALT = 'F8F9FA'
C_WHITE = 'FFFFFF'

THIN_BORDER = Border(
    left=Side(style='thin', color='DEE2E6'),
    right=Side(style='thin', color='DEE2E6'),
    top=Side(style='thin', color='DEE2E6'),
    bottom=Side(style='thin', color='DEE2E6'),
)


def _hfill(color): return PatternFill('solid', fgColor=color)
def _font(bold=False, color='000000', size=10): return Font(bold=bold, color=color, size=size, name='Calibri')
def _align(h='left', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_header_row(ws, row_idx, values, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill = _hfill(bg)
        cell.font = _font(bold=True, color=fg, size=size)
        cell.alignment = _align('center')
        cell.border = THIN_BORDER


def _write_data_row(ws, row_idx, values, alternate=False):
    bg = C_ROW_ALT if alternate else C_WHITE
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill = _hfill(bg)
        cell.alignment = _align('left')
        cell.border = THIN_BORDER
        cell.font = _font(size=9)


def _auto_width(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        length = min_w
        for cell in col_cells:
            try:
                if cell.value:
                    length = max(length, min(len(str(cell.value)) + 2, max_w))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


# sheet builders
def _sheet_summary(wb, analysis):
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    # title block
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"AutoInsight - {analysis.get('filename', 'analysis')}"
    title_cell.font = _font(bold=True, color=C_ACCENT, size=16)
    title_cell.alignment = _align('left', 'center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:F2')
    ws['A2'].value = 'Automated Exploratory Data Analysis Report'
    ws['A2'].font = _font(color='888888', size=10)
    ws['A2'].alignment = _align('left', 'center')
    ws.row_dimensions[2].height = 18

    # kpi boxes (row 4+)
    kpis = [
        ('File', analysis.get('filename', '-')),
        ('Rows', analysis.get('shape', [0, 0])[0]),
        ('Columns', analysis.get('shape', [0, 0])[1]),
        ('Total Cells', analysis.get('missing_summary', {}).get('total_cells', 0)),
        ('Missing Cells', analysis.get('missing_summary', {}).get('total_missing', 0)),
        ('Missing %', f"{analysis.get('missing_summary', {}).get('missing_percentage', 0)}%"),
        ('Cols with Missing', analysis.get('missing_summary', {}).get('columns_with_missing', 0)),
    ]

    _write_header_row(ws, 4, ['Metric', 'Value'])
    for i, (label, val) in enumerate(kpis, 5):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True, size=9)
        ws.cell(row=i, column=1).fill = _hfill('F1F3FF')
        ws.cell(row=i, column=1).border = THIN_BORDER
        cell = ws.cell(row=i, column=2, value=val)
        cell.font = _font(size=9)
        cell.border = THIN_BORDER
        cell.alignment = _align('left')
    _auto_width(ws)


def _sheet_statistics(wb, analysis):
    ws = wb.create_sheet('Statistics')
    ws.sheet_view.showGridLines = False

    columns = analysis.get('columns', [])
    stats = analysis.get('stats', {})
    dtypes = analysis.get('dtypes', {})

    if not stats:
        ws['A1'].value = 'No statistics available.'
        return

    # collect all metric keys
    all_metrics = set()
    for col in columns:
        if col in stats:
            all_metrics.update(stats[col].keys())
    metric_order = ['count', 'mean', 'std', 'min', '25%', '50%', '75%', 'max',
                    'skew', 'kurt', 'unique_values', 'most_common', 'sample_values']
    metric_order += [m for m in all_metrics if m not in metric_order]

    header = ['Metric'] + [f"{col}\n({dtypes.get(col,'')})" for col in columns if col in stats]
    _write_header_row(ws, 1, header)
    ws.row_dimensions[1].height = 36

    for ri, metric in enumerate(metric_order, 2):
        ws.cell(row=ri, column=1, value=metric).font = _font(bold=True, size=9)
        ws.cell(row=ri, column=1).fill = _hfill('F1F3FF')
        ws.cell(row=ri, column=1).border = THIN_BORDER
        ws.cell(row=ri, column=1).alignment = _align('left')
        ci = 2
        for col in columns:
            if col not in stats:
                continue
            val = stats[col].get(metric)
            if isinstance(val, list):
                val = ', '.join(str(v) for v in val[:5])
            elif val is None:
                val = '-'
            elif isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = _hfill(C_ROW_ALT if ri % 2 == 0 else C_WHITE)
            cell.border = THIN_BORDER
            cell.font = _font(size=9)
            cell.alignment = _align('center')
            ci += 1

    _auto_width(ws, min_w=12)


def _sheet_missing(wb, analysis):
    ws = wb.create_sheet('Missing Values')
    ws.sheet_view.showGridLines = False

    ms = analysis.get('missing_summary', {})
    mv = analysis.get('missing_values', {})
    md = analysis.get('missing_details', {})
    rows = analysis.get('shape', [1, 1])[0] or 1

    # summary block
    _write_header_row(ws, 1, ['Missing Value Summary'], bg=C_ACCENT)
    ws.merge_cells('A1:G1')
    summary_data = [
        ('Total Cells', ms.get('total_cells', 0)),
        ('Total Missing', ms.get('total_missing', 0)),
        ('Overall Missing %', f"{ms.get('missing_percentage', 0)}%"),
        ('Columns Affected', ms.get('columns_with_missing', 0)),
    ]
    for i, (k, v) in enumerate(summary_data, 2):
        ws.cell(row=i, column=1, value=k).font = _font(bold=True, size=9)
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2, value=v).font = _font(size=9)
        ws.cell(row=i, column=2).border = THIN_BORDER

    # per-column breakdown
    headers = ['Column', 'Missing Count', 'Missing %', 'NaN/None', 'Empty Strings',
               'Null Keywords', 'Placeholders', 'Whitespace']
    _write_header_row(ws, 7, headers)

    for ri, col in enumerate(analysis.get('columns', []), 8):
        n_miss = mv.get(col, 0)
        pct = round(n_miss / rows * 100, 2)
        det = md.get(col, {})
        row_vals = [
            col, n_miss, f'{pct}%',
            det.get('pandas_na', 0),
            det.get('empty_strings', 0),
            det.get('null_keywords', 0),
            det.get('placeholders', 0),
            det.get('whitespace_only', 0),
        ]
        alt = ri % 2 == 0
        _write_data_row(ws, ri, row_vals, alternate=alt)

        # color-coding the missing count cell
        miss_cell = ws.cell(row=ri, column=2)
        if pct > 30:
            miss_cell.fill = _hfill('FECACA')
        elif pct > 5:
            miss_cell.fill = _hfill('FEF3C7')
        elif pct > 0:
            miss_cell.fill = _hfill('D1FAE5')

    _auto_width(ws)


def _sheet_outliers(wb, analysis):
    ws = wb.create_sheet('Outliers')
    ws.sheet_view.showGridLines = False

    outliers = analysis.get('outliers', {})
    if not outliers:
        ws['A1'].value = 'No outlier data available (numeric columns required).'
        return

    _write_header_row(ws, 1, ['Column', 'Outlier Count', 'Outlier %', 'Lower Fence', 'Upper Fence', 'Status'])

    for ri, (col, data) in enumerate(outliers.items(), 2):
        count = data.get('count', 0)
        pct = data.get('pct', 0)
        status = 'High' if pct > 10 else ('Moderate' if count > 0 else 'Clean')
        row_vals = [col, count, f'{pct}%', data.get('lower', '-'), data.get('upper', '-'), status]
        _write_data_row(ws, ri, row_vals, alternate=ri % 2 == 0)

        status_cell = ws.cell(row=ri, column=6)
        if pct > 10:
            status_cell.fill = _hfill('FECACA')
        elif count > 0:
            status_cell.fill = _hfill('FEF3C7')
        else:
            status_cell.fill = _hfill('D1FAE5')

    _auto_width(ws)


def _sheet_preview(wb, analysis):
    ws = wb.create_sheet('Data Preview')
    ws.sheet_view.showGridLines = False

    preview = analysis.get('data_preview', {})
    cols = preview.get('columns', [])
    rows = preview.get('rows', [])

    if not cols:
        ws['A1'].value = 'No preview available.'
        return

    _write_header_row(ws, 1, ['#'] + cols)

    for ri, row in enumerate(rows, 2):
        _write_data_row(ws, ri, [ri - 1] + [str(v) if v is not None else '' for v in row], alternate=ri % 2 == 0)

    _auto_width(ws)


# public api
def generate_excel(analysis: dict) -> bytes:
    wb = Workbook()

    _sheet_summary(wb, analysis)
    _sheet_statistics(wb, analysis)
    _sheet_missing(wb, analysis)
    _sheet_outliers(wb, analysis)
    _sheet_preview(wb, analysis)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()